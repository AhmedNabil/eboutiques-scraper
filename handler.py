from itertools import product
from logging import info, warn
from math import e, inf, prod
import re
from tkinter import XView
from bs4 import BeautifulSoup
import requests
import csv
from pathlib import Path
from const import cats
import xlsxwriter
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import uuid


def url_handler(baseUrl, slug, pagination_slug, page):
    url = baseUrl + slug
    if page > 1:
        url = (
            baseUrl + slug + "/" + "?category=" + pagination_slug + "&page=" + str(page)
        )
    return url


def soup_response(url):
    try:
        session = requests.Session()
        session.cookies.update({"__hs_opt_out": "yes"})
        session.get(url)
        response = requests.get(url)
        if response.status_code == 403:
            headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
            response = requests.get(url, headers=headers).text
            soup = BeautifulSoup(response, "lxml")
            return soup
        else:
            soup = BeautifulSoup(response.text, "lxml")
            return soup
    except Exception as e:
        print(e)


def next_page_checker(url):
    soup = soup_response(url)
    if soup.find("a", class_="next_page") is None:
        return False
    else:
        return True


def product_list_handler(url):
    soup = soup_response(url)
    products = soup.find_all("div", class_="product-tile")
    return products


def product_handler(product):
    text_wrapper = product.find("div", class_="text-wrapper")
    product_name = text_wrapper.find("div", class_="product-name").text
    product_company = text_wrapper.find("div", class_="product-company").text
    product_img = product.find("img")["src"]
    product_url = product.find("a")["href"]
    return {
        "product_name": product_name,
        "product_company": product_company,
        "product_img": product_img,
        "product_url": product_url,
    }


def product_details_handler(url):
    soup = soup_response(url)
    product_name = soup.find("h2", class_="product-name").text
    label_information = fetch_label_information(soup)
    product_ingredient_concerns = fetch_product_ingredient_concerns(soup)
    return {
        "product_name": product_name,
        "label_information": label_information,
        "product_ingredient_concerns": product_ingredient_concerns,
    }


def fetch_product_ingredient_concerns(soup):
    ingredient_list = []
    if soup.find("section", class_="product-concerns-and-info") is None:
        return None

    ingredient_list = (
        soup.find("section", class_="product-concerns-and-info")
        .find("table", class_="table-ingredient-concerns")
        .find("tbody")
        .find_all("tr", class_="ingredient-overview-tr")
    )

    ingredient_more_info_list = (
        soup.find("section", class_="product-concerns-and-info")
        .find("table", class_="table-ingredient-concerns")
        .find("tbody")
        .find_all("tr", class_="ingredient-more-info-wrapper")
    )

    product_ingredient_concern = []
    for idx, ingredient in enumerate(ingredient_list):
        ingredient_concern = {}
        ingredient_concern["ingredient_name"] = (
            ingredient.find("td", class_="td-ingredient")
            .find("div", class_="td-ingredient-interior")
            .text.strip()
        )
        for heading in ingredient_more_info_list[idx].find_all(["td"]):
            if "CONCERNS" in heading.text:
                ingredient_concern["ingredient_concerns"] = heading.find_next("td").text
            if "FUNCTION(S)" in heading.text:
                ingredient_concern["ingredient_function"] = heading.find_next("td").text
        product_ingredient_concern.append(ingredient_concern)
    return product_ingredient_concern


def fetch_label_information(soup):
    info_list = []
    if soup.find("section", attrs={"id": "label-information"}) is None:
        return None
    info_list = soup.find("section", attrs={"id": "label-information"}).find_all(
        ["h2", "p"]
    )
    for heading in info_list:
        label_information = {}
        if (
            heading.name == "h2"
            and "Ingredients from packaging" in heading.text.strip()
        ):
            for p in heading.find_next_siblings("p"):
                label_information["ingredients_from_packaging"] = p.text.strip()

        if heading.name == "h2" and "Directions from packaging" in heading.text.strip():
            for p in heading.find_next_siblings("p"):
                label_information["directions_from_packaging"] = p.text.strip()

        if heading.name == "h2" and "Warnings from packaging" in heading.text.strip():
            for p in heading.find_next_siblings("p"):
                label_information["warnings_from_packaging"] = p.text.strip()

        return label_information


def img_download(url, id):
    img = requests.get(url)
    with open(f"imgs/{id}.png", "wb") as file:
        session = requests.Session()
        session.cookies.update({"__hs_opt_out": "yes"})
        session.get(url)
        response = requests.get(url)
        if response.status_code == 403:
            headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
            response = requests.get(url, headers=headers, stream=True)
        if not response.ok:
            print(response)
        for block in response.iter_content(1024):
            if not block:
                break
            file.write(block)


def xlsx_file(FILE_PATH):
    if not FILE_PATH.exists():
        with open(FILE_PATH, "w", newline="") as products_xlsx:
            workbook = xlsxwriter.Workbook("products.xlsx")
            products_worksheet = workbook.add_worksheet("Products")
            ingredients_worksheet = workbook.add_worksheet("Ingredients")
            bold = workbook.add_format({"bold": True})

            products_worksheet.set_column("A:A", 20)
            products_worksheet.set_column("B:B", 40)
            products_worksheet.set_column("C:C", 40)
            products_worksheet.set_column("D:D", 40)
            products_worksheet.set_column("E:E", 20)
            products_worksheet.set_column("F:F", 20)
            products_worksheet.set_column("G:G", 20)
            products_worksheet.set_column("H:H", 100)
            products_worksheet.set_column("I:I", 100)
            products_worksheet.set_column("J:J", 100)

            products_worksheet.write("A1", "ID", bold)
            products_worksheet.write("B1", "Name", bold)
            products_worksheet.write("C1", "Image", bold)
            products_worksheet.write("D1", "Product URL", bold)
            products_worksheet.write("E1", "Main Category", bold)
            products_worksheet.write("F1", "Parent Category", bold)
            products_worksheet.write("G1", "Sub Category", bold)
            products_worksheet.write("H1", "Ingredients From Packaging", bold)
            products_worksheet.write("I1", "Directions From Packaging", bold)
            products_worksheet.write("J1", "Warnings From Packaging", bold)

            ingredients_worksheet.set_column("A:A", 20)
            ingredients_worksheet.set_column("B:B", 80)
            ingredients_worksheet.set_column("C:C", 20)
            ingredients_worksheet.set_column("D:D", 150)
            ingredients_worksheet.set_column("E:E", 150)

            ingredients_worksheet.write("A1", "ID", bold)
            ingredients_worksheet.write("B1", "Ingredient Name", bold)
            ingredients_worksheet.write("C1", "Product ID", bold)
            ingredients_worksheet.write("D1", "FUNCTION(S)", bold)
            ingredients_worksheet.write("E1", "CONCERNS", bold)
            print("Workbook Created")
            workbook.close()


def get_sheets(FILE_PATH):
    wb = load_workbook(FILE_PATH)
    sheets = wb.sheetnames
    print(wb.worksheets)
    return sheets


def write_product_to_xlsx(FILE_PATH, product):
    wb = load_workbook(FILE_PATH)
    ws = wb["Products"]
    img_download(product["product_img"], product["id"])
    img = openpyxl.drawing.image.Image(f"imgs/{product['id']}.png")
    row = ws.max_row + 1
    img.anchor = f"C{row}"
    ws.cell(row=row, column=1, value=product["id"])
    ws.cell(row=row, column=2, value=product["product_name"])
    ws.cell(row=row, column=3, value=product["product_img"])
    # ws.add_image(img, f"C{row}")
    ws.cell(row=row, column=4, value=product["product_url"])
    ws.cell(row=row, column=5, value=product["main_category"])
    ws.cell(row=row, column=6, value=product["parent_category"])
    ws.cell(row=row, column=7, value=product["sub_category"])
    ws.cell(row=row, column=8, value=product["ingredient_from_packaging"])
    ws.cell(row=row, column=9, value=product["directions_from_packaging"])
    ws.cell(row=row, column=10, value=product["warnings_from_packaging"])
    wb.save(FILE_PATH)


def write_ingredient_to_xlsx(FILE_PATH, ingredient):
    wb = load_workbook(FILE_PATH)
    ws = wb["Ingredients"]
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=uuid.uuid4().hex)
    ws.cell(row=row, column=2, value=ingredient["ingredient_name"])
    ws.cell(row=row, column=3, value=ingredient["product_id"])
    ws.cell(row=row, column=4, value=ingredient["ingredient_function"])
    ws.cell(row=row, column=5, value=ingredient["ingredient_concerns"])
    wb.save(FILE_PATH)
